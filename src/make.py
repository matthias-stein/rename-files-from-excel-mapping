#!/usr/bin/env python3
"""
Read file_srce/name_deck.xlsx and copy renamed files as
defined there.
"""
from openpyxl import load_workbook

from dataclasses import dataclass, field
from pathlib import Path
from shutil import copy2


def main():
    """Control the program flow"""
    xlsx = Xlsx()
    xlsx.open_file()
    xlsx.open_deck()
    xlsx.copy_file()


class Xlsx:
    """Read Excel file mapping and process files"""

    def __init__(self):
        self.book = None
        self.data = []
        self.path = Path(__file__).parent.parent

    def open_file(self) -> None:
        """Verify existence of and open Excel file."""
        xlsx = self.path / 'file_srce' / 'name_deck.xlsx'
        try:
            self.book = load_workbook(filename=xlsx)
        except FileNotFoundError:
            print(xlsx, 'does not exist!')
            exit()

    def open_deck(self) -> None:
        """Verify existence of and read sheet named 'deck'."""
        try:
            assert 'deck' in self.book.sheetnames
        except AssertionError:
            print('No sheet named "deck" in workbook!')
            exit()

    def copy_file(self) -> None:
        """Create a class for each file and call its copy method."""
        shet = self.book['deck']
        for irow in range(2, shet.max_row + 1, 1):
            srce = shet.cell(row=irow, column=1).value
            dest = shet.cell(row=irow, column=2).value.replace(" ", "_")
            this = File(srce, dest)
            this.copy_file()

@dataclass
class File:
    """Each instance represents one file to be renamed."""
    # Class internal variables
    dest_path: Path = field(init=False)
    srce_path: Path = field(init=False)

    # Excel data
    srce: str
    dest: str

    def __post_init__(self) -> None:
        base = Path(__file__).parent.parent
        self.srce_path = base / 'file_srce' / self.srce
        self.dest_path = base / 'file_dest' / self.dest

    def copy_file(self) -> None:
        """Copy file after verifying existence."""
        print()
        print('Processing', self.srce, '>', self.dest)
        if self.test_srce() is False:
            return
        if self.test_dest() is False:
            return
        copy2(self.srce_path, self.dest_path)

    def test_srce(self) -> bool:
        """Verify source file exists."""
        try:
            assert self.srce_path.is_file()
            return True
        except AssertionError:
            print('   ', self.srce_path, 'does not exist!')
            return False

    def test_dest(self) -> bool:
        """Verify dest dir exists but dest file does not exist."""
        try:
            assert self.dest_path.parent.is_dir()
        except AssertionError:
            print('   ', self.dest_path.parent, 'does not exist!')
            return False

        try:
            assert self.dest_path.is_file() is False
        except AssertionError:
            print('   ', self.dest_path, 'already exists!')
            return False

        return True


if __name__ == '__main__':
    main()
